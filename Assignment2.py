import click
import openpyxl
from copy import copy

@click.command()
@click.option('--capitalize / --no-capitalize',default=False,help='all string data will be capitalized during copy into destination')
@click.option('--preservestyle / --no-preservestyle',default=False,help='the cell styles will be copied')
@click.argument('src',nargs=1,required=True)
@click.argument('dest',nargs=1,required=True)
def copyexcel(src,dest,capitalize,preservestyle):
    '''Copies data from one excel to another'''
    srcwb=openpyxl.load_workbook(filename=src)
    destwb=openpyxl.Workbook()
    destwb.remove(destwb.active)
    for sheet in srcwb.sheetnames:
        dest_sheet=destwb.create_sheet(sheet)
        for row in srcwb[sheet].rows:
            for cell in row:
                new_cell=dest_sheet.cell(row=cell.row,column=cell.col_idx)
                if capitalize:
                    new_cell.value=cell.value.upper()
                else:
                    new_cell.value=cell.value
                if preservestyle:
                    new_cell.alignment=copy(cell.alignment)
                    new_cell.border=copy(cell.border)
                    new_cell.fill=copy(cell.fill)
                    new_cell.style=copy(cell.style)
                    new_cell.font=copy(cell.font)
                    new_cell.protection=copy(cell.protection)
    destwb.save(dest)

if __name__=="__main__":
    copyexcel()