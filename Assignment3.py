import click
import openpyxl
from bs4 import BeautifulSoup

@click.command()
@click.argument('src',nargs=1,required=True)
@click.argument('dest',nargs=1,required=True)
def extract_scores(src,dest):
    '''Extract table from a webpage (.html) into an excel (.xlsx) sheet'''
    src_file=BeautifulSoup(open(src,"r"),'html.parser')
    rows=src_file.find_all("tr")
    dest_excel=openpyxl.Workbook()
    sheet=dest_excel.active
    row_id=1
    col_id=1
    heads=src_file.find_all("th")
    for head in heads[1:]:
        sheet.cell(row=row_id,column=col_id,value=head.string)
        col_id+=1
    for row in rows:
        col_id=1
        data=row.find_all("td")
        data=data[1:]
        for cell in data:
            sheet.cell(row=row_id,column=col_id,value=cell.string)
            col_id+=1
        row_id += 1
    dest_excel.save(dest)

if __name__=="__main__":
    extract_scores()